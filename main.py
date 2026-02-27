import subprocess
import time
import pyautogui
import pyperclip
from openpyxl import load_workbook
from PIL import ImageGrab, ImageOps
import pytesseract
import re
import os
import sys
import logging
from datetime import datetime
import configparser

# Configure pytesseract to find tesseract
# This is required for pytesseract to work properly
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Global configuration
EXE_PATH = r"C:\InfoVotantes\InfoVotantes.exe"

def get_app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def load_screen_region_from_config():
    """
    Load screen region coordinates from config.ini if it exists.
    Returns tuple (x, y, width, height) or None if not found.
    """
    config_file = os.path.join(get_app_dir(), "config.ini")
    
    if not os.path.exists(config_file):
        return None
    
    try:
        config = configparser.ConfigParser()
        config.read(config_file)
        
        if 'SCREEN_REGION' in config:
            x = int(config['SCREEN_REGION']['x'])
            y = int(config['SCREEN_REGION']['y'])
            width = int(config['SCREEN_REGION']['width'])
            height = int(config['SCREEN_REGION']['height'])
            return (x, y, width, height)
    except Exception as e:
        print(f"Warning: Could not read config.ini: {e}")
    
    return None

# Configure logging - output to both console and file
LOG_FILE_PATH = os.path.join(get_app_dir(), f"infovotantes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

# Create logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Create formatters
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%H:%M:%S')

# Console handler
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(formatter)

# File handler
file_handler = logging.FileHandler(LOG_FILE_PATH, encoding='utf-8')
file_handler.setLevel(logging.INFO)
file_handler.setFormatter(formatter)

# Add handlers to logger
logger.addHandler(console_handler)
logger.addHandler(file_handler)

EXCEL_FILE_PATH = os.path.join(get_app_dir(), "InfoVotantes.xlsx")
CEDULA = "1117512408"

# Region coordinates for the result area (x, y, width, height)
# Try to load from config.ini first, otherwise use default
config_region = load_screen_region_from_config()
if config_region:
    RESULT_REGION = config_region
    logger.info(f"Loaded screen region from config.ini: {RESULT_REGION}")
else:
    RESULT_REGION = (932, 334, 305, 314)
    logger.info(f"Using default screen region: {RESULT_REGION}")
    logger.info("Run getwindow.py to configure for this computer's screen")

# Extra pixels around the result area to avoid clipping
REGION_PADDING = 0

# Disable pyautogui fail-safe
pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.5


def run_exe():
    """
    Launches the InfoVotantes.exe application.
    Returns the subprocess object for later interaction if needed.
    """
    try:
        logger.info("Launching InfoVotantes application...")
        process = subprocess.Popen(EXE_PATH)
        time.sleep(2)
        return process
    except FileNotFoundError:
        logger.error(f"Executable not found at {EXE_PATH}")
        return None
    except Exception as e:
        logger.error(f"Error launching application: {e}")
        return None


def extract_text_from_screen(x=0, y=0, width=None, height=None):
    """
    Takes a screenshot of the screen and extracts text using OCR.
    If width/height not specified, captures the entire screen.
    """
    try:
        if width is None or height is None:
            # Capture full screen
            screenshot = ImageGrab.grab()
        else:
            # Capture specific region
            screenshot = ImageGrab.grab(bbox=(x, y, x + width, y + height))
        # Improve contrast for more stable OCR across screens
        preprocessed = ImageOps.grayscale(screenshot)
        preprocessed = ImageOps.autocontrast(preprocessed)

        # Use pytesseract to extract text from the screenshot
        text = pytesseract.image_to_string(preprocessed)
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting text from screen: {e}")
        return ""


def parse_voting_info(raw_text):
    """
    Parses the raw OCR text and formats it into structured voting information.
    Returns a dictionary with the parsed data.
    """
    lines = [line.strip() for line in raw_text.split('\n') if line.strip()]

    # Post-processing: Fix common OCR errors
    if len(lines) >= 3:
        # Fix Line 2 (Zona): "0" often interpreted as "i¢)"
        lines[2] = lines[2].replace('i¢)', '0').replace('ic)', '0').replace('i©)', '0')
        # Also handle other common "0" misreads
        lines[2] = lines[2].replace('O', '0')  # Capital O to zero
        # Keep only digits in Zona (e.g., "3(" or "3{" -> "3")
        zona_digits = re.findall(r"\d+", lines[2])
        if zona_digits:
            lines[2] = zona_digits[0]

    result = {
        'Departamento': '',
        'Municipio': '',
        'Zona': '',
        'Puesto': '',
        'Mesa': '',
        'Direccion': ''
    }

    if len(lines) < 4:
        return result

    # Line 0 -> Departamento
    result['Departamento'] = lines[0]
    
    # Line 1 -> Municipio
    result['Municipio'] = lines[1]
    
    # Line 2 -> Zona (as is)
    result['Zona'] = lines[2]
    
    # Line 3 -> Puesto (first part)
    result['Puesto'] = lines[3]

    if len(lines) < 5:
        return result

    # Line 4: Check if starts with letter or number
    if lines[4] and lines[4][0].isalpha():
        # Line 4 is continuation of Puesto (starts with letter)
        puesto_parts = [result['Puesto'], lines[4]]
        next_index = 5

        # Line 5 can also be continuation of Puesto if it starts with a letter
        if next_index < len(lines) and lines[next_index][0].isalpha():
            puesto_parts.append(lines[next_index])
            next_index += 1

        result['Puesto'] = ' '.join(puesto_parts)

        # Next line after Puesto is Mesa
        if next_index < len(lines):
            result['Mesa'] = lines[next_index]
            # Clean Mesa: keep only digits
            mesa_digits = re.findall(r"\d+", result['Mesa'])
            if mesa_digits:
                result['Mesa'] = mesa_digits[0]
            next_index += 1

            # Everything after Mesa is Direccion
            if next_index < len(lines):
                result['Direccion'] = ' '.join(lines[next_index:])
    else:
        # Line 4 is Mesa (starts with number)
        result['Mesa'] = lines[4]
        # Clean Mesa: keep only digits
        mesa_digits = re.findall(r"\d+", result['Mesa'])
        if mesa_digits:
            result['Mesa'] = mesa_digits[0]

        # Line 5 and onwards is Direccion
        if len(lines) >= 6:
            result['Direccion'] = ' '.join(lines[5:])

    return result


def capture_result_text_with_retry(region, retries=2, delay_seconds=1.5):
    """
    Captures OCR text from a region with retries to allow UI to fully render.
    Returns the best (longest) OCR text observed.
    """
    padded_region = (
        max(0, region[0] - REGION_PADDING),
        max(0, region[1] - REGION_PADDING),
        region[2] + (REGION_PADDING * 2),
        region[3] + (REGION_PADDING * 2)
    )
    best_text = ""
    for attempt in range(retries + 1):
        text = extract_text_from_screen(
            x=padded_region[0],
            y=padded_region[1],
            width=padded_region[2],
            height=padded_region[3]
        )
        if len(text) > len(best_text):
            best_text = text
        if attempt < retries:
            time.sleep(delay_seconds)
    return best_text


def format_voting_info(parsed_data):
    """
    Formats the parsed voting information for display.
    """
    output = []
    output.append(f"Departamento: {parsed_data['Departamento']}")
    output.append(f"Municipio: {parsed_data['Municipio']}")
    output.append(f"Zona: {parsed_data['Zona']}")
    output.append(f"Puesto: {parsed_data['Puesto']}")
    output.append(f"Mesa: {parsed_data['Mesa']}")
    output.append(f"Direccion: {parsed_data['Direccion']}")
    return '\n'.join(output)


def enter_cedula_and_search(cedula):
    """
    Enters the cedula into the text box and presses Enter to search.
    Then reads the results using OCR and returns to the cedula input page.
    """
    try:
        # Copy to clipboard and paste
        pyperclip.copy(cedula)
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(1)

        # Press Enter to search
        pyautogui.press('enter')
        time.sleep(4)  # Wait for search results to load

        # Extract text from results using OCR
        result_text = capture_result_text_with_retry(RESULT_REGION, retries=2, delay_seconds=1.5)
        
        # Parse and print formatted output
        parsed_result = parse_voting_info(result_text)
        formatted_result = format_voting_info(parsed_result)
        logger.info(f"\nResult:\n{formatted_result}\n")

        # Press Enter again to go back to cedula input page
        pyautogui.press('enter')
        time.sleep(2)  # Wait for page to load

        return parsed_result
    except Exception as e:
        logger.error(f"Error processing cedula {cedula}: {e}")
        return None


def read_cedulas_from_excel():
    """
    Reads cedulas from Column A of the Excel file starting from row 2.
    Returns a list of cedula values.
    """
    try:
        logger.info(f"Reading cedulas from {EXCEL_FILE_PATH}")
        workbook = load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook.active

        cedulas = []
        row = 2
        while True:
            cell_value = worksheet[f'A{row}'].value
            if cell_value is None:
                break
            cedulas.append(str(cell_value).strip())
            row += 1

        logger.info(f"Found {len(cedulas)} cedulas")
        return cedulas
    except FileNotFoundError:
        logger.error(f"Excel file not found at {EXCEL_FILE_PATH}")
        return []
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        return []


def write_voting_data_to_excel(row_number, voting_data):
    """
    Writes parsed voting data to Excel columns B-G for the specified row.
    B: Departamento, C: Municipio, D: Zona, E: Puesto, F: Mesa, G: Direccion
    """
    try:
        workbook = load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook.active
        
        worksheet[f'B{row_number}'] = voting_data['Departamento']
        worksheet[f'C{row_number}'] = voting_data['Municipio']
        worksheet[f'D{row_number}'] = voting_data['Zona']
        worksheet[f'E{row_number}'] = voting_data['Puesto']
        worksheet[f'F{row_number}'] = voting_data['Mesa']
        worksheet[f'G{row_number}'] = voting_data['Direccion']
        
        workbook.save(EXCEL_FILE_PATH)
        return True
    except Exception as e:
        logger.error(f"Error writing to Excel row {row_number}: {e}")
        return False


def main():
    """
    Main function to orchestrate the automation workflow.
    """
    logger.info("=" * 50)
    logger.info("InfoVotantes Automation Started")
    logger.info(f"Log file: {LOG_FILE_PATH}")
    logger.info("=" * 50)

    # Step 1: Read cedulas from Excel
    cedulas = read_cedulas_from_excel()
    if not cedulas:
        logger.error("No cedulas found. Exiting.")
        return

    # Step 2: Run the exe
    process = run_exe()

    if process:
        # Wait for application to be ready
        time.sleep(3)

        # Step 3: Process each cedula
        total_cedulas = len(cedulas)
        for index, cedula in enumerate(cedulas, 1):
            logger.info(f"\n{'='*50}")
            logger.info(f"Processing cedula {index} of {total_cedulas}: {cedula}")
            logger.info(f"{'='*50}")
            voting_data = enter_cedula_and_search(cedula)
            
            # Write data to Excel (row starts at 2, so row_number = index + 1)
            if voting_data:
                write_voting_data_to_excel(index + 1, voting_data)
                formatted_output = format_voting_info(voting_data)
                logger.info(f"\nResult:\n{formatted_output}\n")
            else:
                logger.warning(f"No data returned for cedula {cedula}")

            time.sleep(2)  # Wait between searches

        logger.info("=" * 50)
        logger.info(f"All {total_cedulas} cedulas processed successfully")
        logger.info("=" * 50)
    else:
        logger.error("Failed to launch application. Exiting.")


if __name__ == '__main__':
    main()